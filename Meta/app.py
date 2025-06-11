#importa os modulos necessarios, para o flask funcionar
from flask import Flask, render_template,request,redirect
#importando as funcoes da biblioteca openpyxl p/ criar e manipular um arquivo exccel
from openpyxl import Workbook,load_workbook
#Biblioteca p/ verificar a existência de um arq.exccel
import os

#Criando a aplicação:
app = Flask(__name__)
#Definindo o nome da planilha exccel
ARQUIVO = 'vendas.xlsx'

if not os.path.exists(ARQUIVO):
    wb = Workbook() #CRIANDO UM ARQ.EXCCEL
    ws = wb.active #Selecionando a planilha ativa do projeto
    #Criando um cabeçalho p/ a planilha
    ws.append(['Nome',  'Vendas', 'Meta'])
    #salvando o arquivo
    wb.save(ARQUIVO)

#22/05
#rota principal do site (formulario cadastro de vendas)
@app.route('/')
def index():
    return render_template('index.html')
    #função que deve ser executada, ao ser requisitado a rota '/'
    #abrir a pagina principal, neste caso é o index.html

#Rota que processa os dados do formulario e salva no exccel
@app.route('/salvar',methods=['POST'])
def salvar():
    #CAPTURA OS DADOS DE CADA UMA DAS CAIXAS DO FORMULARIO E ATRIBUI PARA AS VARIAVEIS
    nome = request.form['nome']
    vendas = float(request.form['vendas'])
    meta = float(request.form['meta'])
    #ABRINDO O ARQUIVO EXCCEL
    wb = load_workbook(ARQUIVO)
    #SELECIONANDO A PLANILHA ATIVA - 1ª ABA POR PADRÃO
    ws = wb.active

    #adiciona uma nova linha como lista com as informações do formulário
    ws.append([nome,vendas,meta])

    #salvando o arquivo exccel
    wb.save(ARQUIVO)

    #redirecionando a rota p/ analisar (onde abrirá uma nova pagina) passando por parâmetro o nome do funcionario.
    return  redirect('/analisar?nome='+nome)
#Rota p/ a tela resultado... analisando se o funcionario bateu a meta
@app.route('/analisar')
def analisar():
    #pegar o parametro do nome do funcionario enviado como parâmetro pra url
    nome_param = request.args.get('nome')

    wb = load_workbook(ARQUIVO)
    ws = wb.active

    #Loop for: percorre todas as linhas da planilha a partir da 2ª linha, pois a 1ª linha é cabeçalho
    for linha in ws.iter_rows(min_row=2,values_only=True):
        nome, vendas, meta = linha
        #a variavel linha sempre receberá três valores, ex:
        #linha = ('Ana', 45, 50) na linha de codigo acima, estamos atribuindo cada elemento da linha a uma variavel na sequencia
        # as variaveis nome, vendas, meta recebem Ana, 45, 50 respectivamente. nome tecnico = desempacotamento

        #Verifica se o nome atual da linha é o mesmo enviado na URL
        if nome == nome_param:
            meta_batida = vendas >= meta
            #Se a meta for batida bonus recebe o resultado do calculo de 15% do valor das vendas, caso contrario recebe 0
            bonus = round(vendas * 0.15, 2) if meta_batida else 0
            #Exibir a tela resultado com as informações dos calculos
            return render_template('resultado.html',nome = nome, meta_batida = meta_batida, bonus = bonus)
    return 'Funcionário não encontrado'

    #Rota que mostra pag. do historico de todos os funcionarios cadastrados
@app.route('/historico')
def historico():
    wb = load_workbook(ARQUIVO)
    ws = wb.active
    #Converte os dados da planilha (a partir da 2ª linha) em uma tupla
    dados = list(ws.iter_rows(min_row=2, values_only=True))
    return render_template('historico.html', dados = dados)
#Iniciando o Flask no modo desenvolvedor Debug
if __name__ == '__main__':
    app.run(debug=True)
